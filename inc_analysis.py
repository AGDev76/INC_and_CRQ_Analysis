from openpyxl import load_workbook, workbook
from openpyxl.formatting.rule import ColorScaleRule
from fuzzywuzzy import fuzz, process
import sys
import getopt


def main(argv):
    # ***** Provide usage methods to the user *****
    inputfile = ''
    outputfile = ''
    try:
        opts, args = getopt.getopt(argv, "hi:o:", ["ifile=", "ofile="])
    except getopt.GetoptError:
        print('Usage: python3 yourfile.py -i <inputfile> -o <outputfile>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('Usage: python3 yourfile.py -i <inputfile> -o <outputfile>')
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg

    if inputfile == '':
        print("Usage: python3 yourfile.py -i <inputfile> -o <outputfile>")
        return
    else:
        # Before beginning it provides a suggestion to the user in order to view data properly
        print("-----------------------")
        print("Análisis de Incidentes")
        print("-----------------------", "\n\n")
        ready = input("No olvides marcar la Col P, ir a Format-Cells-Alignment y marcar Wrap text. Pulsá 1 y <ENTER> para continuar: ")
        if ready == "1":
            print("\nProcesando ... esto puede demorar varias horas ...")
            workbook = load_workbook(filename=inputfile)
            sheet = workbook.active
            server_names = []
            res_counter = 2
            det_counter = 2

            # Process txt file with a list of words to validate and creates a list with results
            try:
                with open("keywords.txt", "r") as keywords:
                    for row in keywords:
                        data = row.strip("\n")
                        server_names.append(data)
            except Exception as e:
                print(e)

            # If Title cells don't exist, it creates them with the following titles
            if sheet["R1"].value == None and sheet["S1"].value == None and sheet["T1"].value == None and sheet["U1"].value == None:
                sheet["R1"] = "Res_Server_Name"
                sheet["S1"] = "Res_Word_Match"
                sheet["T1"] = "Det_Server_Name"
                sheet["U1"] = "Det_Word_Match"
                try:
                    workbook.save(filename=outputfile)
                except Exception as e:
                    print(e)

            # Iterate through rows and validates only column 15. Results are being added on Col R and S
            for value in sheet.iter_rows(min_row=2, min_col=15, max_col=15, values_only=True):
                if value[0] != None:
                    result = process.extractOne(
                        value[0], server_names, scorer=fuzz.partial_ratio)
                    percent = result[1]
                    if percent > 80: 
                        sheet["R" + str(res_counter)] = result[0]
                        sheet["S" + str(res_counter)] = result[1]
                        try:
                            workbook.save(filename=outputfile)
                        except Exception as e:
                            print(e)
                    res_counter += 1
                else:
                    res_counter += 1

            # Iterate through rows and validates only column 16. Results are being added on Col T and U
            for value in sheet.iter_rows(min_row=2, min_col=16, max_col=16, values_only=True):
                if value[0] != None:
                    result = process.extractOne(
                        value[0], server_names, scorer=fuzz.partial_ratio)
                    percent = result[1]
                    if percent > 80: 
                        sheet["T" + str(det_counter)] = result[0]
                        sheet["U" + str(det_counter)] = result[1]
                        try:
                            workbook.save(filename=outputfile)
                        except Exception as e:
                            print(e)
                    det_counter += 1
                else:
                    det_counter += 1

            # Paint percentage cells with different colors according to result
            color_scale_rule = ColorScaleRule(start_type="num",
                                            start_value=80,
                                            start_color="00FF0000",
                                            mid_type="num",
                                            mid_value=90,
                                            mid_color="00FFFF00",
                                            end_type="num",
                                            end_value=100,
                                            end_color="0000FF00")

            sheet.conditional_formatting.add("S2:S35000", color_scale_rule)
            sheet.conditional_formatting.add("U2:U35000", color_scale_rule)
            workbook.save(filename=outputfile)
        else:
            print("Opción no válida")
            return

if __name__ == "__main__":
    main(sys.argv[1:])
