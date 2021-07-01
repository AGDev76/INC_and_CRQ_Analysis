from openpyxl import load_workbook, workbook
from openpyxl.formatting.rule import ColorScaleRule
from fuzzywuzzy import fuzz, process
import sys
import getopt


def main(argv):
    inputfile = ''
    outputfile = ''
    try:
        opts, args = getopt.getopt(argv, "hi:o:", ["ifile=", "ofile="])
    except getopt.GetoptError:
        print('Usage: python3 yourfile.py -i <inputfile> -o <outputfile>')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('Usage: python3 yourfile.py -i <inputfile> -o <outputfile>')
            sys.exit()
        elif opt in ("-i", "--ifile"):
            inputfile = arg
        elif opt in ("-o", "--ofile"):
            outputfile = arg

    if inputfile == '':
        print("Usage: python3 yourfile.py -i <inputfile> -o <outputfile>")
        return
    else:
        print("----------------------------")
        print("Categorización de Incidentes")
        print("----------------------------", "\n\n")
        print("Input file: " + inputfile)
        print("Output file: " + outputfile)
        ready = input(
            "Cuando todo esté listo ingresá 1 y <ENTER> para continuar: ")
        if ready == "1":
            print("\nProcesando ... esto puede demorar varias horas ...")
            workbook = load_workbook(filename=inputfile)
            sheets = workbook.sheetnames
            sheet1 = workbook[sheets[0]]
            sheet2 = workbook[sheets[1]]
            sheet3 = workbook[sheets[2]]
            list_key = []
            list_value = []
            res_counter = 2
            det_counter = 2

            for row in sheet2.values:
                for value in row:
                    if value != None:
                        list_key.append(value)

            for row in sheet3.values:
                for value in row:
                    if value != None:
                        list_value.append(value)

            for value in sheet1.iter_rows(min_row=2, max_col=1, values_only=True):
                if value != None:
                    result = process.extractOne(
                        str(value), list_key, scorer=fuzz.partial_ratio)
                    percent = result[1]
                    if percent == 100:
                        sheet1["C" + str(res_counter)
                               ] = list_value[list_key.index(result[0])]
                        try:
                            workbook.save(filename=outputfile)
                        except Exception as e:
                            print(e)
                    res_counter += 1
                else:
                    res_counter += 1

            for value in sheet1.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
                if value != None:
                    result = process.extractOne(
                        str(value), list_key, scorer=fuzz.partial_ratio)
                    percent = result[1]
                    if percent == 100:
                        sheet1["D" + str(det_counter)
                               ] = list_value[list_key.index(result[0])]
                        try:
                            workbook.save(filename=outputfile)
                        except Exception as e:
                            print(e)
                    det_counter += 1
                else:
                    det_counter += 1

            # color_scale_rule = ColorScaleRule(start_type="num",
            #                                   start_value=80,
            #                                   start_color="00FF0000",
            #                                   mid_type="num",
            #                                   mid_value=90,
            #                                   mid_color="00FFFF00",
            #                                   end_type="num",
            #                                   end_value=100,
            #                                   end_color="0000FF00")

            # cells.conditional_formatting.add("S2:S25000", color_scale_rule)
            # cells.conditional_formatting.add("U2:U25000", color_scale_rule)
            # workbook.save(filename=outputfile)
        else:
            print("Opción no válida")
            return


if __name__ == "__main__":
    main(sys.argv[1:])
