from openpyxl import load_workbook, workbook
from openpyxl.formatting.rule import ColorScaleRule
from fuzzywuzzy import fuzz, process

workbook = load_workbook(filename="data2.xlsx")
sheet = workbook.active
server_names = []
counter = 2

try:
    with open("keywords.txt", "r") as keywords:    
        for row in keywords:
            data = row.strip("\n")
            server_names.append(data)        
except Exception as e:
    print(e)

if sheet["CP1"].value == None and sheet["CQ1"].value == None:
    sheet["CP1"] = "Server_Name"
    sheet["CQ1"] = "Word_Match"
    try:
        workbook.save(filename="data2.xlsx")
    except Exception as e:
        print(e)

for value in sheet.iter_rows(min_row=2, min_col=31, max_col=31,values_only=True):
    if value[0] != None:
        result = process.extractOne(value[0], server_names, scorer=fuzz.partial_ratio)
        percent = result[1]
        if percent > 80 and sheet["CP" + str(counter)] != None and sheet["CQ" + str(counter)] != result[1] != None:
            sheet["CP" + str(counter)] = result[0]    
            sheet["CQ" + str(counter)] = result[1]
            try:
                workbook.save(filename="data2.xlsx")        
            except Exception as e:
                print(e)            
        counter += 1
    else:
        counter += 1

color_scale_rule = ColorScaleRule(start_type="num",
                                    start_value=80,
                                    start_color="00FF0000",
                                    mid_type="num",
                                    mid_value=90,
                                    mid_color="00FFFF00",
                                    end_type="num",
                                    end_value=100,
                                    end_color="0000FF00")

sheet.conditional_formatting.add("CQ2:CQ500", color_scale_rule)
workbook.save(filename="data2.xlsx")
